"""床面積表をExcel（.xlsx）またはCSVに書き出す。"""
from __future__ import annotations

import csv
import os
from typing import Iterable

from .rooms import Room
from .table import build_table, is_summary_row

XLSX_SUFFIXES = (".xlsx", ".xlsm")
CSV_SUFFIXES = (".csv",)

DEFAULT_SHEET_TITLE = "床面積表"


class ExportError(Exception):
    """出力に失敗した（openpyxlが無い、拡張子が不明、書き込めない等）。"""


def openpyxl_available() -> bool:
    """openpyxlが使えるか。FreeCAD同梱のPythonには入っていないことがあります。"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def write_area_table(
    rooms: Iterable[Room],
    path: str,
    *,
    sheet_title: str = DEFAULT_SHEET_TITLE,
) -> str:
    """床面積表を `path` に書き出し、書き出したパスを返す。

    拡張子が `.xlsx` / `.xlsm` ならExcel、`.csv` ならCSV（Excelでそのまま
    開けるBOM付きUTF-8）です。
    """
    headers, body = build_table(rooms)
    suffix = os.path.splitext(path)[1].lower()

    if suffix in XLSX_SUFFIXES:
        _write_xlsx(headers, body, path, sheet_title)
    elif suffix in CSV_SUFFIXES:
        _write_csv(headers, body, path)
    else:
        raise ExportError(
            f"対応していない拡張子です: {suffix or '(なし)'}"
            "（.xlsx か .csv を指定してください）"
        )
    return path


def _write_csv(headers, body, path: str) -> None:
    try:
        # Excelで開いても文字化けしないようBOM付きUTF-8にします。
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in body:
                writer.writerow(["" if cell is None else cell for cell in row])
    except OSError as exc:
        raise ExportError(f"CSVを書き出せませんでした: {exc}") from exc


def _write_xlsx(headers, body, path: str, sheet_title: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise ExportError(
            "openpyxlが見つかりません。`pip install openpyxl` で入れるか、"
            "出力先の拡張子を .csv にしてください"
            "（FreeCAD同梱のPythonにはopenpyxlが無いことがあります）。"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    for row in body:
        ws.append(row)

    area_col = len(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="B0B0B0")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center")

    for offset, row in enumerate(body):
        excel_row = offset + 2
        bold = is_summary_row(row)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=excel_row, column=col)
            if bold:
                cell.font = header_font
            if col == area_col:
                cell.number_format = "0.00"
        if bold:
            ws.cell(row=excel_row, column=1).border = Border(top=thin)

    widths = [max(12, min(40, _display_width(headers[i], body, i) + 2))
              for i in range(len(headers))]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
    except OSError as exc:
        raise ExportError(
            f"Excelファイルを保存できませんでした: {exc}"
            "（同じファイルをExcelで開いたままにしていませんか）"
        ) from exc


def _display_width(header, body, index: int) -> int:
    """列幅の目安。全角文字は2文字分として数えます。"""
    def width(value) -> int:
        text = "" if value is None else str(value)
        return sum(2 if ord(ch) > 0x2E80 else 1 for ch in text)

    return max(width(value) for value in [header] + [row[index] for row in body])
